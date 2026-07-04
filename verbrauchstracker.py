#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verbrauchs-Tracker – Strom / Gas / Wasser
==========================================
Plattformübergreifende Desktop-Anwendung (Windows & Linux) zur Verwaltung
von Zählerständen mit grafischer Auswertung und Excel-Export.

Abhängigkeiten:  customtkinter, matplotlib, openpyxl
    pip install customtkinter matplotlib openpyxl

Start:  python verbrauchstracker.py
"""

import os
import sys
import json
import datetime

try:
    import customtkinter as ctk
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except Exception as e:  # pragma: no cover
    print("Fehlende GUI-Bibliothek. Bitte installieren:  pip install customtkinter")
    raise

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates

# ---------------------------------------------------------------------------
# Konfiguration der drei Verbrauchsarten
# ---------------------------------------------------------------------------
UTILITIES = {
    "strom":  {"label": "Strom",  "einheit": "kWh", "farbe": "#f2b705"},
    "gas":    {"label": "Gas",    "einheit": "m\u00b3", "farbe": "#e8590c"},
    "wasser": {"label": "Wasser", "einheit": "m\u00b3", "farbe": "#1c7ed6"},
}

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "verbrauchsdaten.json")

DATE_FORMATS = ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y")


# ---------------------------------------------------------------------------
# Datenhaltung
# ---------------------------------------------------------------------------
def parse_date(text):
    """Datum aus mehreren Formaten einlesen -> date-Objekt oder None."""
    text = (text or "").strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(text):
    """Zahl mit Komma oder Punkt einlesen -> float oder None."""
    text = (text or "").strip().replace(" ", "")
    if text == "":
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def load_data(path=DATA_FILE):
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            raw = json.load(fh)
    except (json.JSONDecodeError, OSError):
        return []
    result = []
    for r in raw:
        dt = parse_date(r.get("datum"))
        if dt is None:
            continue
        result.append({
            "datum": dt,
            "strom": r.get("strom"),
            "gas": r.get("gas"),
            "wasser": r.get("wasser"),
        })
    result.sort(key=lambda x: x["datum"])
    return result


def save_data(readings, path=DATA_FILE):
    payload = [{
        "datum": r["datum"].isoformat(),
        "strom": r["strom"],
        "gas": r["gas"],
        "wasser": r["wasser"],
    } for r in readings]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


def compute_consumption(readings):
    """Aus aufeinanderfolgenden Zählerständen den Verbrauch je Zeitraum
    berechnen. Ein negativer Wert (Zählerwechsel) wird als None markiert,
    damit die Diagramme keinen unsinnigen Ausschlag zeigen."""
    periods = []
    for i in range(1, len(readings)):
        prev, cur = readings[i - 1], readings[i]
        tage = (cur["datum"] - prev["datum"]).days
        row = {"datum": cur["datum"], "tage": tage}
        for key in UTILITIES:
            a, b = prev.get(key), cur.get(key)
            if a is None or b is None:
                row[key] = None
                row[key + "_wechsel"] = False
            else:
                diff = b - a
                if diff < 0:
                    row[key] = None          # Zählerwechsel
                    row[key + "_wechsel"] = True
                else:
                    row[key] = diff
                    row[key + "_wechsel"] = False
        periods.append(row)
    return periods


# ---------------------------------------------------------------------------
# Excel-Export
# ---------------------------------------------------------------------------
def export_excel(readings, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()

    # --- Blatt 1: Zählerstände -------------------------------------------
    ws = wb.active
    ws.title = "Z\u00e4hlerst\u00e4nde"
    header = ["Datum", "Strom (kWh)", "Gas (m\u00b3)", "Wasser (m\u00b3)"]
    ws.append(header)
    for r in readings:
        ws.append([r["datum"], r["strom"], r["gas"], r["wasser"]])
    ws.column_dimensions["A"].number_format = "DD.MM.YYYY"

    # --- Blatt 2: Verbrauch je Zeitraum ----------------------------------
    ws2 = wb.create_sheet("Verbrauch")
    ws2.append(["Datum", "Tage", "Strom (kWh)", "Gas (m\u00b3)", "Wasser (m\u00b3)"])
    periods = compute_consumption(readings)
    for p in periods:
        ws2.append([p["datum"], p["tage"], p["strom"], p["gas"], p["wasser"]])

    # --- Formatierung beider Blätter -------------------------------------
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(bottom=thin, right=thin)
    for sheet in (ws, ws2):
        for cell in sheet[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center")
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            sheet.column_dimensions[col[0].column_letter].width = width + 3
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = "DD.MM.YYYY"

    # --- Diagramm auf dem Verbrauchsblatt --------------------------------
    n = len(periods)
    if n > 0:
        colmap = {"strom": 3, "gas": 4, "wasser": 5}
        for key, col in colmap.items():
            chart = LineChart()
            chart.title = f"{UTILITIES[key]['label']} – Verbrauch je Zeitraum"
            chart.y_axis.title = UTILITIES[key]["einheit"]
            chart.x_axis.title = "Datum"
            chart.height = 7
            chart.width = 18
            data_ref = Reference(ws2, min_col=col, min_row=1, max_row=n + 1)
            cats = Reference(ws2, min_col=1, min_row=2, max_row=n + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            anchor_row = 2 + (col - 3) * 16
            ws2.add_chart(chart, f"H{anchor_row}")

    wb.save(path)


# ---------------------------------------------------------------------------
# Hauptfenster
# ---------------------------------------------------------------------------
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")

        self.title("Verbrauchs-Tracker  \u2013  Strom / Gas / Wasser")
        self.geometry("1180x720")
        self.minsize(980, 620)

        self.readings = load_data()
        self.current_util = "strom"

        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_main()
        self.refresh_all()

    # -- linke Seitenleiste ------------------------------------------------
    def _build_sidebar(self):
        side = ctk.CTkFrame(self, width=280, corner_radius=0)
        side.grid(row=0, column=0, sticky="nsew")
        side.grid_propagate(False)

        ctk.CTkLabel(side, text="Neuer Z\u00e4hlerstand",
                     font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(22, 12), padx=20)

        self.entries = {}
        today = datetime.date.today().strftime("%d.%m.%Y")

        self._add_field(side, "datum", "Datum", today)
        for key, meta in UTILITIES.items():
            self._add_field(side, key, f"{meta['label']} ({meta['einheit']})", "")

        ctk.CTkButton(side, text="\u2795  Hinzuf\u00fcgen",
                      command=self.add_reading, height=40).pack(pady=(16, 6), padx=20, fill="x")
        ctk.CTkButton(side, text="\U0001F5D1  Ausgew\u00e4hlten l\u00f6schen",
                      command=self.delete_selected, height=36,
                      fg_color="#b02a37", hover_color="#8b2029").pack(pady=6, padx=20, fill="x")

        ctk.CTkFrame(side, height=2, fg_color="#888").pack(fill="x", padx=20, pady=18)

        ctk.CTkButton(side, text="\U0001F4C4  Export nach Excel",
                      command=self.do_export, height=40,
                      fg_color="#2f9e44", hover_color="#26803a").pack(pady=6, padx=20, fill="x")

        self.status = ctk.CTkLabel(side, text="", wraplength=240,
                                   font=ctk.CTkFont(size=12), text_color="#7d8590")
        self.status.pack(pady=(14, 10), padx=20)

        ctk.CTkLabel(side, text="Darstellung", font=ctk.CTkFont(size=12)).pack(pady=(6, 2))
        ctk.CTkOptionMenu(side, values=["System", "Hell", "Dunkel"],
                          command=self._change_mode, width=200).pack(pady=(0, 16))

    def _add_field(self, parent, key, label, default):
        ctk.CTkLabel(parent, text=label, anchor="w").pack(fill="x", padx=20, pady=(8, 0))
        e = ctk.CTkEntry(parent, height=34)
        e.pack(fill="x", padx=20)
        if default:
            e.insert(0, default)
        self.entries[key] = e

    def _change_mode(self, choice):
        ctk.set_appearance_mode({"System": "system", "Hell": "light", "Dunkel": "dark"}[choice])
        self.after(60, self.draw_chart)

    # -- rechter Hauptbereich ---------------------------------------------
    def _build_main(self):
        main = ctk.CTkFrame(self, corner_radius=0)
        main.grid(row=0, column=1, sticky="nsew")
        main.grid_rowconfigure(1, weight=1)
        main.grid_rowconfigure(2, weight=1)
        main.grid_columnconfigure(0, weight=1)

        # Umschalter Strom / Gas / Wasser
        self.util_switch = ctk.CTkSegmentedButton(
            main, values=[UTILITIES[k]["label"] for k in UTILITIES],
            command=self._switch_util, height=38,
            font=ctk.CTkFont(size=14, weight="bold"))
        self.util_switch.set(UTILITIES[self.current_util]["label"])
        self.util_switch.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 6))

        # Diagramm
        chart_frame = ctk.CTkFrame(main)
        chart_frame.grid(row=1, column=0, sticky="nsew", padx=16, pady=6)
        chart_frame.grid_rowconfigure(0, weight=1)
        chart_frame.grid_columnconfigure(0, weight=1)
        self.fig = Figure(figsize=(7, 3.4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")

        # Tabelle
        table_frame = ctk.CTkFrame(main)
        table_frame.grid(row=2, column=0, sticky="nsew", padx=16, pady=(6, 14))
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        cols = ("datum", "strom", "gas", "wasser",
                "v_strom", "v_gas", "v_wasser", "tage")
        heads = ("Datum", "Strom\nStand", "Gas\nStand", "Wasser\nStand",
                 "Strom\nVerbr.", "Gas\nVerbr.", "Wasser\nVerbr.", "Tage")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings",
                                 selectmode="browse")
        for c, h in zip(cols, heads):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=95, anchor="center")
        self.tree.column("datum", width=100)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.grid(row=0, column=1, sticky="ns")
        self._style_tree()

    def _style_tree(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Treeview", rowheight=26, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))

    # -- Aktionen ----------------------------------------------------------
    def _switch_util(self, label):
        for k, m in UTILITIES.items():
            if m["label"] == label:
                self.current_util = k
                break
        self.draw_chart()

    def add_reading(self):
        dt = parse_date(self.entries["datum"].get())
        if dt is None:
            messagebox.showerror("Ung\u00fcltiges Datum",
                                 "Bitte Datum im Format TT.MM.JJJJ eingeben.")
            return
        vals = {}
        for key in UTILITIES:
            vals[key] = parse_number(self.entries[key].get())
        if all(vals[k] is None for k in UTILITIES):
            messagebox.showerror("Keine Werte",
                                 "Bitte mindestens einen Z\u00e4hlerstand eingeben.")
            return
        # bestehendes Datum ersetzen?
        for r in self.readings:
            if r["datum"] == dt:
                if not messagebox.askyesno("Datum vorhanden",
                        f"F\u00fcr den {dt.strftime('%d.%m.%Y')} existiert bereits "
                        "ein Eintrag. \u00dcberschreiben?"):
                    return
                r.update(vals)
                break
        else:
            self.readings.append({"datum": dt, **vals})
        self.readings.sort(key=lambda x: x["datum"])
        save_data(self.readings)
        for key in UTILITIES:
            self.entries[key].delete(0, "end")
        self.set_status(f"Eintrag {dt.strftime('%d.%m.%Y')} gespeichert.")
        self.refresh_all()

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            self.set_status("Kein Eintrag ausgew\u00e4hlt.")
            return
        idx = int(sel[0])
        r = self.readings[idx]
        if messagebox.askyesno("L\u00f6schen",
                f"Eintrag {r['datum'].strftime('%d.%m.%Y')} wirklich l\u00f6schen?"):
            del self.readings[idx]
            save_data(self.readings)
            self.set_status("Eintrag gel\u00f6scht.")
            self.refresh_all()

    def do_export(self):
        if not self.readings:
            messagebox.showinfo("Export", "Keine Daten zum Exportieren.")
            return
        default = f"Verbrauch_{datetime.date.today().isoformat()}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile=default,
            filetypes=[("Excel-Datei", "*.xlsx")])
        if not path:
            return
        try:
            export_excel(self.readings, path)
            self.set_status(f"Exportiert nach:\n{os.path.basename(path)}")
            messagebox.showinfo("Export erfolgreich",
                                f"Datei gespeichert:\n{path}")
        except Exception as e:
            messagebox.showerror("Export fehlgeschlagen", str(e))

    def set_status(self, text):
        self.status.configure(text=text)

    # -- Anzeige aktualisieren --------------------------------------------
    def refresh_all(self):
        self.fill_table()
        self.draw_chart()

    def fill_table(self):
        self.tree.delete(*self.tree.get_children())
        periods = {p["datum"]: p for p in compute_consumption(self.readings)}

        def fmt(v, dec=1):
            return "" if v is None else f"{v:,.{dec}f}".replace(",", "\u00a0")

        for idx, r in enumerate(self.readings):
            p = periods.get(r["datum"], {})
            def vv(key):
                if key not in p:
                    return ""
                if p.get(key + "_wechsel"):
                    return "Wechsel"
                return fmt(p.get(key))
            self.tree.insert("", "end", iid=str(idx), values=(
                r["datum"].strftime("%d.%m.%Y"),
                fmt(r["strom"]), fmt(r["gas"]), fmt(r["wasser"]),
                vv("strom"), vv("gas"), vv("wasser"),
                p.get("tage", ""),
            ))
        children = self.tree.get_children()
        if children:
            self.tree.see(children[-1])

    def draw_chart(self):
        key = self.current_util
        meta = UTILITIES[key]
        periods = compute_consumption(self.readings)
        xs = [p["datum"] for p in periods if p.get(key) is not None]
        ys = [p[key] for p in periods if p.get(key) is not None]

        dark = ctk.get_appearance_mode() == "Dark"
        bg = "#1a1a1a" if dark else "#ffffff"
        fg = "#e6e6e6" if dark else "#222222"
        grid = "#444444" if dark else "#dddddd"

        self.fig.set_facecolor(bg)
        self.ax.clear()
        self.ax.set_facecolor(bg)
        if xs:
            self.ax.plot(xs, ys, marker="o", markersize=4, linewidth=2,
                         color=meta["farbe"])
            self.ax.fill_between(xs, ys, color=meta["farbe"], alpha=0.12)
        self.ax.set_title(f"{meta['label']} \u2013 Verbrauch je Ableseperiode",
                          color=fg, fontsize=12, fontweight="bold")
        self.ax.set_ylabel(meta["einheit"], color=fg)
        self.ax.tick_params(colors=fg, labelsize=9)
        for spine in self.ax.spines.values():
            spine.set_color(grid)
        self.ax.grid(True, color=grid, linewidth=0.6, alpha=0.7)
        self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%y"))
        self.fig.autofmt_xdate(rotation=35)
        self.fig.tight_layout()
        self.canvas.draw()


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
